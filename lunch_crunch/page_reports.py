#   LunchCrunch: A Python desktop app to manage food ordering
#
#   Copyright (C) 2025  Fabian Sauer
#
#   This program is free software: you can redistribute it and/or modify
#   it under the terms of the GNU Affero General Public License as published
#   by the Free Software Foundation, either version 3 of the License, or
#   (at your option) any later version.
#
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU Affero General Public License for more details.
#
#   You should have received a copy of the GNU Affero General Public License
#   along with this program.  If not, see <https://www.gnu.org/licenses/>.

"""Reports page - monthly meal counts per child."""

import calendar
from datetime import date
from itertools import groupby

import io
import openpyxl
from openpyxl.styles import Font
from nicegui import ui

from lunch_crunch.db import get_db
from lunch_crunch.common import weekdays_of_month, get_children, get_closing_days, get_setting, header
from lunch_crunch.filter import month_and_group_filter

@ui.page("/reports")
def reports_page() -> None:
    """Route "/reports" - monthly meal count summary and per-child breakdown."""
    header()

    today = date.today()
    current = {"year": today.year, "month": today.month, "group": None}

    def _load_month(year: int, month: int) -> tuple:
        """Return (children, days, child_counts, grand_total) for the given month."""
        _, last = calendar.monthrange(year, month)
        month_str = f"{year}-{month:02d}"
        with get_db() as conn:
            children       = get_children(conn, f"{month_str}-01", f"{month_str}-{last:02d}", current["group"])
            closed         = get_closing_days(conn, year, month)
            days           = [d for d in weekdays_of_month(year, month) if d.isoformat() not in closed]
            absent         = {
                (r["child_id"], r["date"])
                for r in conn.execute(
                    "SELECT child_id, date FROM absence WHERE date LIKE ?", (f"{month_str}-%",)
                ).fetchall()
            }
            holiday_absent = {
                (r["child_id"], r["date"])
                for r in conn.execute(
                    "SELECT child_id, date FROM holiday_absence WHERE date LIKE ?", (f"{month_str}-%",)
                ).fetchall()
            }
            price_str  = get_setting(conn, "price_per_meal", "")
            price      = float(price_str) if price_str else 0.0
        child_counts = []
        grand_total  = 0
        for child in children:
            meals = sum(
                1 for d in days
                if (child["id"], d.isoformat()) not in absent
                and (child["id"], d.isoformat()) not in holiday_absent
            )
            child_counts.append((child, meals))
            grand_total += meals
        return children, days, child_counts, grand_total, price

    def has_data(year: int, month: int) -> bool:
        month_str = f"{year}-{month:02d}"
        _, last = calendar.monthrange(year, month)
        with get_db() as conn:
            return bool(get_children(conn, f"{month_str}-01", f"{month_str}-{last}", current["group"]))

    def export() -> None:
        """Build an .xlsx with the monthly meal summary and download it."""
        year, month = current["year"], current["month"]
        children, days, child_counts, grand_total, price = _load_month(year, month)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = date(year, month, 1).strftime("%B %Y")

        ws.append([date(year, month, 1).strftime("%B %Y")])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])

        bold = Font(bold=True)
        if price:
            ws.append(["", "", "", "Kosten (€)"])
            ws[ws.max_row][3].font = bold
            ws.append(["", "", "p.P.", price])
            ws.append([])
        
        ws.append(["", "Name", "Essen"])
        for cell in ws[ws.max_row]:
            cell.font = bold

        for group_name, group_items in groupby(child_counts, key=lambda x: x[0]["group_name"]):
            ws.append([group_name])
            ws[ws.max_row][0].font = bold
            for child, meals in group_items:
                if price:
                    ws.append(["", child["name"], meals, f"=C{ws.max_row + 1}*D4"])
                else:
                    ws.append(["", child["name"], meals])
        ws.append([])

        last_data_row = ws.max_row - 1
        if price:
            ws.append(["Gesamtsumme", "", f"=SUM(C7:C{ws.max_row})", f"=SUM(D7:D{ws.max_row})"])
        else:
            ws.append(["Gesamtsumme", "", f"=SUM(C4:C{ws.max_row})"])
        for cell in ws[ws.max_row]:
            cell.font = bold

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"Essen_{year}-{month:02d}.xlsx"
        ui.download(buf.getvalue(), filename)
        ui.notify(f"Export gestartet: {filename}", type="positive")

    def rebuild() -> None:
        year, month = current["year"], current["month"]
        children, days, child_counts, grand_total, price = _load_month(year, month)

        is_forecast = (year, month) >= (today.year, today.month)
        download_btn.set_enabled(not is_forecast)

        # -- Summary card ------------------------------------------------------
        summary_card.clear()
        with summary_card:
            with ui.row().classes("items-center gap-2"):
                ui.label("Monatsübersicht").classes("font-medium")
                if is_forecast:
                    ui.badge("Prognose", color="orange").props("outline")
            ui.separator()
            with ui.row().classes("gap-8"):
                with ui.column().classes("gap-0 items-center"):
                    ui.label(str(len(days))).classes("text-3xl font-bold text-blue-600")
                    ui.label("Kindergartentage").classes("text-xs text-gray-500")
                with ui.column().classes("gap-0 items-center"):
                    ui.label(str(len(children))).classes("text-3xl font-bold text-blue-600")
                    ui.label("Kinder").classes("text-xs text-gray-500")
                with ui.column().classes("gap-0 items-center"):
                    ui.label(str(grand_total)).classes("text-3xl font-bold text-green-600")
                    ui.label(
                        "Essen erwartet" if is_forecast else "Essen bestellt"
                    ).classes("text-xs text-gray-500")
                if price:
                    with ui.column().classes("gap-0 items-center"):
                        ui.label(f"{grand_total * price:.2f} €".replace(".", ",")).classes("text-3xl font-bold text-green-600")
                        ui.label("Gesamtkosten").classes("text-xs text-gray-500")

        # -- Per-child card ----------------------------------------------------
        meals_label = "Essen erwartet" if is_forecast else "Essen bestellt"
        children_card.clear()
        with children_card:
            ui.label("Essen pro Kind").classes("font-medium")
            ui.separator()
            if not children:
                ui.label("Keine Kinder registriert.").classes("text-gray-500 text-sm")
            else:
                columns = [
                    {"name": "name",  "label": "Name",        "field": "name",  "align": "left"},
                    {"name": "group", "label": "Gruppe",       "field": "group", "align": "left"},
                    {"name": "meals", "label": meals_label,    "field": "meals", "align": "right"},
                ]
                if price:
                    columns.append({"name": "cost", "label": "Kosten", "field": "cost", "align": "right"})
                ui.table(
                    rows=[
                        {
                            "name":  child["name"],
                            "group": child["group_name"],
                            "meals": meals,
                            "cost":  f"{meals * price:.2f} €".replace(".", ","),
                        }
                        for child, meals in child_counts
                    ],
                    columns=columns,
                ).classes("w-full")

    with ui.column().classes("w-full max-w-3xl mx-auto"):
        ui.label("Berichte").classes("text-2xl font-semibold").style("font-family: Antropos;")
        ui.separator()

        with ui.row(align_items="center"):
            # Month navigation + group filter
            month_and_group_filter(current, update=rebuild, has_data=has_data)
            ui.space()
            download_btn = ui.button("Exportieren", icon="download", on_click=export).props("flat dense")

        summary_card = ui.card().classes("w-full")
        children_card = ui.card().classes("w-full")

    rebuild()
